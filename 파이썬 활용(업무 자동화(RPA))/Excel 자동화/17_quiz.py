# 조건1: 퀴즈2 점수를 10으로 수정
# 조건2: H열에 총점(SUM 이용), I열에 성적 정보 추가
#       총점 90이상 A, 80이상 B, 70이상 C, 나머지 D
# 조건3: 출석이 5 미만인 학생은 총점 상관없이 F

from openpyxl import Workbook

wb = Workbook()
ws = wb.active

# 현재까지 작성된 최종 성적 데이터 넣기
ws.append(("학번", "출석", "퀴즈1", "퀴즈2", "중간고사", "기말고사", "프로젝트"))

scores = [(1, 10, 8, 5, 14, 26, 12),
          (2, 7, 3, 7, 15, 24, 18),
          (3, 9, 5, 8, 8, 12, 4),
          (4, 7, 8, 7, 17, 21, 18),
          (5, 7, 8, 7, 16, 25, 15),
          (6, 3, 5, 8, 8, 17, 0),
          (7, 4, 9, 10, 16, 27, 18),
          (8, 6, 6, 6, 15, 19, 17),
          (9, 10, 10, 9, 19, 30, 19),
          (10, 9, 8, 8, 20, 25, 20)
          ]

for s in scores: # 기존 성적 데이터 넣기
    ws.append(s)
    
# 퀴즈2 점수를 10으로 수정
for idx, cell in enumerate(ws["D"]):
    if idx == 0: # 제목인 경우 skip
        continue
    cell.value = 10

# H열에 총점, I열에 성적 정보 추가
ws["H1"] = "총점"
ws["I1"] = "성적"

for idx, score in enumerate(scores, start=2):
    sum_val = sum(score[1:]) - score[3] + 10 # 총점
    ws.cell(row=idx, column=8).value = "=SUM(B{}:G{})".format(idx, idx)
    
    grade = None # 성적
    if sum_val >= 90:
        grade = "A"
    elif sum_val >= 80:
        grade = "B"
    elif sum_val >= 70:
        grade = "C"
    else:
        grade = "D"
    
    # 출석이 5점 미만이면 F
    if score[1] < 5:
        grade = "F"
        
    ws.cell(row=idx, column=9).value = grade # I열에 성적 정보
    
wb.save("scroes.xlsx")